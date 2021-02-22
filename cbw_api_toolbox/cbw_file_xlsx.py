"""Module used to import export file xlsx remote acesses"""

import datetime
import logging
import openpyxl
import xlsxwriter

from cbw_api_toolbox.cbw_api import CBWApi


class CBWXlsx:
    """Class used to import/export file xlsx remote acesses"""

    def __init__(self, api_url, api_key, secret_key):
        self.client = CBWApi(api_url, api_key, secret_key)

    def import_remote_accesses_xlsx(self, file_xlsx):
        """method to import remote accesses from an xlsx file"""
        if not file_xlsx:
            logging.fatal("No Files xlsx")
            return None

        if not file_xlsx.endswith(".xlsx"):
            logging.fatal("Extension not valid")
            return None

        response = []

        workbook = openpyxl.load_workbook(file_xlsx)
        worksheet = workbook.active
        titles = {}

        for idx, cell in enumerate(worksheet[1]):
            titles[cell.value] = idx

        for row in worksheet.iter_rows(min_row=2):
            try:
                address = row[0].value
                logging.debug("Creating remote access {}".format(address))
                info = {
                    "address": row[titles["HOST"]].value,
                    "port": row[titles["PORT"]].value,
                    "type": row[titles["TYPE"]].value,
                    "login": row[titles["USERNAME"]].value,
                    "password": row[titles["PASSWORD"]].value,
                    "key": row[titles["KEY"]].value,
                    "node_id": row[titles["NODE_ID"]].value,
                    "server_groups": row[titles["SERVER_GROUPS"]].value
                }
                remote_access = self.client.create_remote_access(info)
                response.append(remote_access)
                logging.debug("Done")

            except ValueError:
                logging.fatal("Error format file xlsx::"
                              "HOST, PORT, TYPE, USERNAME,"
                              "PASSWORD, KEY, NODE_ID, SERVER_GROUPS")
                return None
        return response

    def export_remote_accesses_xlsx(self,
                                    file_xlsx="export_remote_accesses_"+str(
                                        datetime.datetime.now())+".xlsx"):
        """Method to export remote access to an xlsx file"""
        if not file_xlsx or file_xlsx == "":
            logging.error("No xlsx file")
            return False

        if not file_xlsx.endswith(".xlsx"):
            logging.error("Extension not valid")
            return False

        remote_accesses = self.client.remote_accesses()

        logging.debug("Creating file xlsx")
        workbook = xlsxwriter.Workbook(file_xlsx)
        worksheet = workbook.add_worksheet(u"remote_accesses")

        worksheet.write(0, 0, "HOST")
        worksheet.write(0, 1, "PORT")
        worksheet.write(0, 2, "TYPE")
        worksheet.write(0, 3, "NODE_ID")
        worksheet.write(0, 4, "SERVER_GROUPS")

        logging.debug("Add remote accesses in file xlsx")

        i = 1
        for remote_access in remote_accesses:
            worksheet.write(i, 0, remote_access.address)
            worksheet.write(i, 1, remote_access.port)
            worksheet.write(i, 2, remote_access.type)
            worksheet.write(i, 3, remote_access.node_id)

            if remote_access.server_id:
                server = self.client.server(str(remote_access.server_id))
                if server.groups:
                    group_name = ""
                    for group in server.groups:
                        group_name += group.name + ","
                    group_name = group_name[:-1]
                    worksheet.write(i, 4, group_name)
            i += 1

        workbook.close()
        logging.debug("Done")
        return True
